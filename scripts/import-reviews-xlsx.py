#!/usr/bin/env python3
"""
Import new Feefo reviews from Excel sheet(s) into Airtable.

Usage (auto mode — recommended):
    AIRTABLE_API_KEY=xxxx python3 scripts/import-reviews-xlsx.py

    Drop .xlsx or .csv files into public/feefo/ (KnowledgeHUB repo root).
    Running the script with no arguments scans that folder, imports every
    .xlsx/.csv file it finds there, and moves each processed file into
    public/feefo/imported/ (renamed with a timestamp prefix) so it won't be
    re-imported next time.

Usage (single file mode):
    AIRTABLE_API_KEY=xxxx python3 scripts/import-reviews-xlsx.py path/to/reviews.xlsx

    Imports just that one file and leaves it where it is (does not move it).

The Excel/CSV file must have a header row with these columns (any order,
case-insensitive). Only Adviser/Review are required; the rest are optional:
    Adviser        -> full adviser name, must match a name in trust-post-owners.json
                       (e.g. "Carl Thorne"). Feefo's native "Customer reference"
                       header is also accepted.
    Customer Name  -> reviewer's name as it should appear on the graphic
    Review         -> the review text. Feefo's native "Service review" header
                       is also accepted.
    Service Rating -> 1-5 star rating (Feefo's "Service rating"/"Rating")
    NPS Rating     -> NPS score (Feefo's "NPS rating"/"NPS")
    Date           -> review date. Feefo's export has no such column, so if
                       omitted the script stamps today's date (the import
                       date) instead.

Rows missing Adviser or Review are skipped. Existing reviews (matched by exact
trimmed Review text, same as the sync task uses) are skipped so you can't
double-import.

This only adds rows to the Airtable "Feefo" table (tblU58wJ0rNFPMiKp in base
appqQv0Xog8yZMwI9) — it does NOT render graphics. The scheduled sync task
picks up new Airtable rows next time it runs and renders them automatically.

Requires: pip install openpyxl requests --break-system-packages
"""
import sys
import os
import csv
import json
import shutil
import datetime
import requests
from openpyxl import load_workbook

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FEEFO_DROP_DIR = os.path.join(REPO_ROOT, "public", "feefo")
FEEFO_IMPORTED_DIR = os.path.join(FEEFO_DROP_DIR, "imported")

BASE_ID = "appqQv0Xog8yZMwI9"
TABLE_ID = "tblU58wJ0rNFPMiKp"
FIELD_ADVISER = "fldHCzSEkK1o4sFno"
FIELD_CUSTOMER = "fld6B1IvhFMPb1exk"
FIELD_REVIEW = "fldkiTpnNhlnS8hmO"
FIELD_RATING = "fld55bC5EsBxDlq6T"   # Service Rating (rating field, 1-5 int)
FIELD_NPS = "fldiCb9Y2p9jaE45D"      # NPS (number)
FIELD_DATE = "fldtrxDQcYIcJHWVp"     # Date (date) — Feefo exports don't include one, so we stamp import date

API_KEY = os.environ.get("AIRTABLE_API_KEY")


def load_rows_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def load_rows_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [tuple(row) for row in csv.reader(f)]


def load_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        rows = load_rows_csv(path)
    else:
        rows = load_rows_xlsx(path)

    if not rows:
        return []
    header = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col(*names):
        for name in names:
            try:
                return header.index(name)
            except ValueError:
                continue
        return None

    # Accepts either our simple headers, or Feefo's native export headers
    # (where confusingly "Customer reference" holds the ADVISER's name, and
    # "Customer name" holds the reviewer's name).
    i_adv = col("adviser", "customer reference")
    i_cust = col("customer name")
    i_rev = col("review", "service review")
    i_rating = col("service rating", "rating")
    i_nps = col("nps rating", "nps")
    i_date = col("date")
    if i_adv is None or i_rev is None:
        sys.exit(f"Header row must include an Adviser/'Customer reference' column and a Review/'Service review' column. Found: {rows[0]}")

    def cell(r, idx):
        return r[idx] if idx is not None and idx < len(r) and r[idx] not in (None, "") else None

    out = []
    for r in rows[1:]:
        adviser = (cell(r, i_adv) or "")
        adviser = str(adviser).strip() if adviser else ""
        review = (cell(r, i_rev) or "")
        review = str(review).strip() if review else ""
        customer = (cell(r, i_cust) or "")
        customer = str(customer).strip() if customer else ""

        rating_raw = cell(r, i_rating)
        try:
            rating = int(float(rating_raw)) if rating_raw is not None else None
        except (ValueError, TypeError):
            rating = None

        nps_raw = cell(r, i_nps)
        try:
            nps = int(float(nps_raw)) if nps_raw is not None else None
        except (ValueError, TypeError):
            nps = None

        date_raw = cell(r, i_date)
        date_val = str(date_raw).strip() if date_raw else None

        if adviser and review:
            out.append({
                "adviser": adviser,
                "customer": customer,
                "review": review,
                "rating": rating,
                "nps": nps,
                "date": date_val,
            })
    return out


def fetch_existing_reviews():
    """Returns dict: adviser name (lower) -> set of trimmed review texts already in Airtable."""
    existing = {}
    url = f"https://api.airtable.com/v0/{BASE_ID}/{TABLE_ID}"
    headers = {"Authorization": f"Bearer {API_KEY}"}
    params = {
        "fields[]": [FIELD_ADVISER, FIELD_REVIEW],
        "pageSize": 100,
    }
    offset = None
    while True:
        p = dict(params)
        if offset:
            p["offset"] = offset
        resp = requests.get(url, headers=headers, params=p)
        resp.raise_for_status()
        data = resp.json()
        for rec in data.get("records", []):
            f = rec.get("fields", {})
            adv = (f.get(FIELD_ADVISER) or "").strip().lower()
            rev = (f.get(FIELD_REVIEW) or "").strip()
            if adv:
                existing.setdefault(adv, set()).add(rev)
        offset = data.get("offset")
        if not offset:
            break
    return existing


def build_fields(r, import_date):
    fields = {
        FIELD_ADVISER: r["adviser"],
        FIELD_CUSTOMER: r["customer"],
        FIELD_REVIEW: r["review"],
    }
    if r.get("rating") is not None:
        fields[FIELD_RATING] = r["rating"]
    if r.get("nps") is not None:
        fields[FIELD_NPS] = r["nps"]
    # Feefo exports don't include a per-review date column, so stamp with
    # the date this row was actually imported (or the sheet's own Date
    # column if present).
    fields[FIELD_DATE] = r.get("date") or import_date
    return fields


def create_records(rows):
    url = f"https://api.airtable.com/v0/{BASE_ID}/{TABLE_ID}"
    headers = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}
    import_date = datetime.date.today().isoformat()
    created = 0
    for i in range(0, len(rows), 10):  # Airtable batch limit = 10
        batch = rows[i:i + 10]
        payload = {
            "records": [
                {"fields": build_fields(r, import_date)}
                for r in batch
            ]
        }
        resp = requests.post(url, headers=headers, json=payload)
        if resp.status_code >= 300:
            print(f"  ERROR on batch {i}: {resp.status_code} {resp.text}")
            continue
        created += len(resp.json().get("records", []))
    return created


def process_file(path, existing, move_after=False):
    rows = load_rows(path)
    print(f"\n{os.path.basename(path)}: read {len(rows)} candidate rows")

    to_create = []
    skipped = 0
    for r in rows:
        adv_key = r["adviser"].lower()
        if r["review"] in existing.get(adv_key, set()):
            skipped += 1
        else:
            to_create.append(r)
            existing.setdefault(adv_key, set()).add(r["review"])  # avoid dup within/between files

    print(f"  New: {len(to_create)}  |  Already in Airtable (skipped): {skipped}")

    created = 0
    if to_create:
        created = create_records(to_create)
        print(f"  Created {created} new record(s) in Airtable Feefo table.")

    if move_after:
        os.makedirs(FEEFO_IMPORTED_DIR, exist_ok=True)
        stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        dest = os.path.join(FEEFO_IMPORTED_DIR, f"{stamp}-{os.path.basename(path)}")
        shutil.move(path, dest)
        print(f"  Moved to {dest}")

    return created


def main():
    if not API_KEY:
        sys.exit("Set AIRTABLE_API_KEY env var first, e.g.:\n  AIRTABLE_API_KEY=xxxx python3 scripts/import-reviews-xlsx.py")

    existing = fetch_existing_reviews()
    total_created = 0

    if len(sys.argv) >= 2:
        # Single-file mode
        total_created += process_file(sys.argv[1], existing, move_after=False)
    else:
        # Auto mode: scan public/feefo/ for .xlsx/.csv files
        if not os.path.isdir(FEEFO_DROP_DIR):
            os.makedirs(FEEFO_DROP_DIR, exist_ok=True)
            print(f"Created drop folder at {FEEFO_DROP_DIR} — drop .xlsx or .csv files there and re-run.")
            return
        files = [
            f for f in sorted(os.listdir(FEEFO_DROP_DIR))
            if f.lower().endswith((".xlsx", ".csv")) and os.path.isfile(os.path.join(FEEFO_DROP_DIR, f))
        ]
        if not files:
            print(f"No .xlsx/.csv files found in {FEEFO_DROP_DIR}. Nothing to do.")
            return
        for fname in files:
            total_created += process_file(os.path.join(FEEFO_DROP_DIR, fname), existing, move_after=True)

    print(f"\nDone. {total_created} new record(s) added to Airtable in total.")
    if total_created:
        print("Run the KnowledgeHUB review-sync task next to render graphics for these.")


if __name__ == "__main__":
    main()
